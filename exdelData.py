import re
import time
import random
import pytest
import openpyxl
from funciones import Funciones_globales
from playwright.sync_api import Page, expect, Playwright, sync_playwright

#Instalar libreria
#pip install openpyxl

#Ejecutar
#pytest exdelData.py -s -v

rutaExcel = "C:/Users/gino.jimenez/Documents/Gino/PersonalProjects/Playwright/Python/Ejemplo/Estudiantes/Ejemplos/data/data.xlsx"

archivoExcel = openpyxl.load_workbook(rutaExcel)
def numero_filas(hoja):
    ac = archivoExcel[hoja]
    return ac.max_row

def dato_columna(hoja,fila,column):
    ac = archivoExcel[hoja]
    col = ac.cell(int(fila),int(column))
    return col.value

print (str(numero_filas("Sheet1"))+" cantidad de filas")
print (dato_columna("Sheet1",1,1))

def test_excel(set_up_parametrizar)-> None:
    page = set_up_parametrizar
    F=Funciones_globales(page)
    F.EscribirTexto("//input[@id='user-name']",dato_columna("Sheet1",1,1),0.1)
    F.EscribirTexto("//input[@id='password']",dato_columna("Sheet1",1,2),0.1)
    F.clickElement("//input[@id='login-button']",0.1)
    F.Esperar(2)